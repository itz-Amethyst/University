import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox
from openpyxl import load_workbook
import os
from main import delete_product_sheet, add_product, edit_product
from utils import excel_file_path
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


class ProductApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Product Management")
        self.root.configure(background="gray")
        self.root.state('zoomed')

        self.current_sheet_index = 0  # current sheet index

        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        # Buttons
        button_frame = tk.Frame(self.root, bg="gray")
        button_frame.grid(row=0, column=0, padx=10, pady=10, sticky=tk.N)

        add_button = tk.Button(button_frame, text="Add", command=self.add_product)
        add_button.pack(fill=tk.X, pady=5)

        update_button = tk.Button(button_frame, text="Update", command=self.update_product)
        update_button.pack(fill=tk.X, pady=5)

        remove_button = tk.Button(button_frame, text="Remove", command=self.delete_product_gui)
        remove_button.pack(fill=tk.X, pady=5)

        # Input Fields
        input_frame = tk.Frame(self.root, bg="gray")
        input_frame.grid(row=0, column=1, padx=10, pady=10, sticky=tk.N)

        tk.Label(input_frame, text="Name:", bg="gray").grid(row=0, column=0, sticky=tk.W)
        self.name_entry = tk.Entry(input_frame)
        self.name_entry.grid(row=0, column=1, pady=5)

        tk.Label(input_frame, text="Price:", bg="gray").grid(row=1, column=0, sticky=tk.W)
        self.price_entry = tk.Entry(input_frame)
        self.price_entry.grid(row=1, column=1, pady=5)

        tk.Label(input_frame, text="Description:", bg="gray").grid(row=2, column=0, sticky=tk.W)
        self.description_entry = tk.Entry(input_frame)
        self.description_entry.grid(row=2, column=1, pady=5)

        tk.Label(input_frame, text="Count:", bg="gray").grid(row=3, column=0, sticky=tk.W)
        self.count_entry = tk.Entry(input_frame)
        self.count_entry.grid(row=3, column=1, pady=5)

        # Separator
        separator = ttk.Separator(input_frame , orient = 'horizontal')
        separator.grid(row = 4 , columnspan = 2 , pady = 10 , sticky = "ew")

        # Date Input Fields
        #? Accepts only date
        tk.Label(input_frame , text = "Start Date (YYYY-MM-DD):" , bg = "gray").grid(row = 5 , column = 0 ,sticky = tk.W)
        self.start_date_entry = tk.Entry(input_frame)
        self.start_date_entry.grid(row = 5 , column = 1 , pady = 5)

        tk.Label(input_frame , text = "End Date (YYYY-MM-DD):" , bg = "gray").grid(row = 6 , column = 0 , sticky = tk.W)
        self.end_date_entry = tk.Entry(input_frame)
        self.end_date_entry.grid(row = 6 , column = 1 , pady = 5)

        update_chart_button = tk.Button(input_frame , text = "Update Chart" , command = self.update_chart_with_filter)
        update_chart_button.grid(row = 7 , columnspan = 2 , pady = 5)

        # Checkbox for applying filter on all sheets
        self.apply_all_sheets_var = tk.BooleanVar()
        apply_all_sheets_checkbox = tk.Checkbutton(input_frame , text = "Apply filter on all sheets" ,variable = self.apply_all_sheets_var)
        apply_all_sheets_checkbox.grid(row = 8 , columnspan = 2 , pady = 5)

        # Treeview
        self.tree = ttk.Treeview(self.root, show='headings')
        self.tree.grid(row=0, column=2, padx=10, pady=10, sticky=tk.NSEW)

        # Navigation Buttons
        nav_frame = tk.Frame(self.root, bg="gray")
        nav_frame.grid(row=1, column=2, pady=10)

        prev_button = tk.Button(nav_frame, text="Previous", command=self.prev_sheet)
        prev_button.grid(row=0, column=0, padx=5)

        next_button = tk.Button(nav_frame, text="Next", command=self.next_sheet)
        next_button.grid(row=0, column=1, padx=5)

        self.root.grid_columnconfigure(2, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        self.chart_frame = tk.Frame(self.root, bg = "gray")
        self.chart_frame.grid(row=0, column=3, padx=10, pady=10, sticky=tk.NSEW)
        self.chart_canvas = None

    def handle_product(self, mode):
        name = self.name_entry.get()
        description = self.description_entry.get()
        stock = self.count_entry.get()
        price = self.price_entry.get()

        # To set the name to last record of the sheet for edit operation
        if not name and mode == "edit":
            ws = self.workbook[self.sheets[self.current_sheet_index]]
            last_row = ws.max_row
            last_record_name = ws.cell(row = last_row , column = 2).value  # Assuming the name is in the 2nd column
            name = last_record_name

        try:
            match mode:
                case "add":
                    if name and description and stock and price:
                        success , msg = add_product(excel_file_path , name , description , int(stock) , int(price))
                        if success:
                            messagebox.showinfo("Success" , "Product operation completed successfully.")
                        else:
                            messagebox.showerror("Error" , msg)
                    else:
                        messagebox.showerror("Error" , "Please fill in all fields for adding a product.")
                case "edit":
                    if name:
                        # Tuple
                        success, msg = edit_product(excel_file_path, self.current_sheet_index , name , description , int(stock) , int(price))
                        if success:
                            messagebox.showinfo("Success" , "Product operation completed successfully.")
                        else:
                            messagebox.showerror("Error" , msg)
                case "delete":
                    if name:
                        success, msg = delete_product_sheet(excel_file_path , name)
                        if success:
                            messagebox.showinfo("Success" , "Product sheet deleted successfully.")
                        else:
                            messagebox.showinfo("Fail" , msg)
                    else:
                        messagebox.showerror("Warning" , "Please enter the product name for deleting.")
        except Exception as e:
            messagebox.showerror("Error" , f"Failed to perform product operation: {e}")

        self.load_data()

    def load_data(self):
        if os.path.exists(excel_file_path):
            self.workbook = load_workbook(excel_file_path)
            self.sheets = self.workbook.sheetnames
            self.display_sheet(self.sheets[self.current_sheet_index])
        else:
            messagebox.showerror("Error", f"Excel file '{excel_file_path}' not found!")

    def display_sheet(self, sheet_name):
        for item in self.tree.get_children():
            self.tree.delete(item)

        ws = self.workbook[sheet_name]
        columns = [cell.value for cell in ws[1]]
        self.tree["columns"] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", "end", values=row)
            data.append(row)

        # print(data)
        self.update_chart(data)

    def update_chart(self, data):
        if self.chart_canvas:
            self.chart_canvas.get_tk_widget().destroy()

        fig, ax = plt.subplots(figsize=(7, 8 ))

        if not data:
            ax.text(0.5, 0.5, 'No data in the specified date range', horizontalalignment='center', verticalalignment='center')
        else:
            prices = [row[4] for row in data]  # 5th column is price
            transaction_dates = [row[0] for row in data]

            ax.plot(transaction_dates, prices, marker='o')
            ax.set_xlabel('Transaction Date')
            ax.set_ylabel('Price')
            ax.set_title('Product Prices Over Time')
            plt.xticks(rotation=45)

        self.chart_canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        self.chart_canvas.draw()
        self.chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        plt.close(fig)  # Close the figure to free up memory

    def update_chart_with_filter(self):

        try:
            start_datetime_str = self.start_date_entry.get()
            end_datetime_str = self.end_date_entry.get()
        except Exception as e:
            messagebox.showinfo("Fail" , "Start_date and end_date must not be empty.")
            print(e)
            return

        try:
            start_date = datetime.strptime(start_datetime_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_datetime_str, "%Y-%m-%d").date()
            if start_date > end_date:
                messagebox.showerror("Error", "Start datetime must be before end datetime.")
                return
        except ValueError:
            messagebox.showerror("Error", "Please enter valid datetimes in the format YYYY-MM-DD HH:MM:SS.")
            return


        if self.apply_all_sheets_var.get():
            # Destroy previous chart canvas
            if self.chart_canvas:
                self.chart_canvas.get_tk_widget().destroy()

            all_data = {}
            for sheet_name in self.sheets:
                ws = self.workbook[sheet_name]
                data = [row for row in ws.iter_rows(min_row = 2 , values_only = True)]
                changes_in_price = []
                prev_price = None
                for row in data:
                    transaction_date = datetime.strptime(row[0] , "%Y-%m-%d %H:%M:%S").date()
                    if start_date <= transaction_date <= end_date:
                        price = row[4]  # 5th is the price
                        if prev_price is not None:
                            change = price - prev_price
                            changes_in_price.append(change)
                        prev_price = price
                all_data[sheet_name] = changes_in_price

            #chart
            fig , ax = plt.subplots(figsize = (7 , 8))
            for sheet_name , changes_in_price in all_data.items():
                ax.plot(changes_in_price , label = sheet_name)
            ax.set_xlabel('Time')
            ax.set_ylabel('Change in Price')
            ax.set_title('Change in Price Over Time for All Sheets')
            # Panel guide on the right side
            ax.legend()
            plt.xticks(rotation = 45)
            self.chart_canvas = FigureCanvasTkAgg(fig , master = self.chart_frame)
            self.chart_canvas.draw()
            self.chart_canvas.get_tk_widget().pack(fill = tk.BOTH , expand = True)
            plt.close(fig)  # freeing up the memory problem
        else:
            ws = self.workbook[self.sheets[self.current_sheet_index]]
            data = [row for row in ws.iter_rows(min_row = 2 , values_only = True)]
            filtered_data = []
            for row in data:
                transaction_date = datetime.strptime(row[0] , "%Y-%m-%d %H:%M:%S").date()
                if start_date <= transaction_date <= end_date:
                    filtered_data.append(row)
            self.update_chart(filtered_data)


    def next_sheet(self):
        if self.current_sheet_index < len(self.sheets) - 2:
            self.current_sheet_index += 1
            self.display_sheet(self.sheets[self.current_sheet_index])

    def prev_sheet(self):
        if self.current_sheet_index > 0:
            self.current_sheet_index -= 1
            self.display_sheet(self.sheets[self.current_sheet_index])

    def add_product(self):
        self.handle_product(mode="add")

    def update_product(self):
        self.handle_product(mode="edit")

    def delete_product_gui(self):
        self.handle_product(mode="delete")


if __name__ == "__main__":
    root = tk.Tk()
    app = ProductApp(root)
    root.mainloop()

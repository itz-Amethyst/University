import os
from typing import Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from utils import data_folder, excel_file_path


def create_excel_file() -> None:
    """Create an Excel file in given path."""
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)

    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active

        # Apply header styles and data validations
        apply_header_styles(ws)
        apply_data_validation(ws)

        wb.save(excel_file_path)
        print(f"Created new Excel file: {excel_file_path}")
    else:
        print(f"Excel file '{excel_file_path}' already exists.")

def apply_header_styles(ws) -> None:
    """Apply styles to the header row."""
    header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        #! Requires a researching
        # cell.protection = Protection(locked=True)
        cell.fill = header_fill

    #!f
    # ws.protection.sheet = True
def apply_data_validation(ws) -> None:
    """Apply data validation to the sheets """

    #! Formula1 : less than 255 char
    dv_text = DataValidation(type="textLength", operator="lessThan", formula1="255", showErrorMessage=True)
    dv_text.error = "Please enter a valid text."
    dv_text.errorTitle = "Invalid Text"
    ws.add_data_validation(dv_text)
    dv_text.add(f"B2:B1000")  # Name
    dv_text.add(f"C2:C1000")  # Description

    dv_int = DataValidation(type="whole", operator="greaterThan", formula1=0, showErrorMessage=True)
    dv_int.error = "Please enter a valid integer."
    dv_int.errorTitle = "Invalid Integer"
    ws.add_data_validation(dv_int)
    dv_int.add(f"D2:D1000")  # Stock
    dv_int.add(f"E2:E1000")  # price

    dv_date = DataValidation(type="date", formula1="1900-01-01", showErrorMessage=True)
    dv_date.error = "Please enter a valid date."
    dv_date.errorTitle = "Invalid Date"
    ws.add_data_validation(dv_date)
    dv_date.add(f"A2:A1000")


def prepare_workbook(excel_file: str):
    """Load the workbook or create if it doesn't exist."""
    if not os.path.exists(excel_file):
        create_excel_file()
    return load_workbook(excel_file)

def save_changes(wb, excel_file: str) -> None:
    try:
        wb.save(excel_file)
        print('Saved Successfully')
    except Exception as e:
        print(f"Failed to save the file: {e}")
        print("There is a chance that this file is in use.")


def validate_sheet_exists(file_name: str , sheet_name: str, flag: bool = False) -> Union[bool , str]:
    """Validate if a sheet exists in the given workbook."""
    if not os.path.exists(file_name):
        msg = f"Excel file '{file_name}' does not exist."
        print(msg)
        return False , msg
    wb = prepare_workbook(file_name)
    # For add operation
    if flag:
        if sheet_name in wb.sheetnames:
            msg = f"Product sheet '{sheet_name}' already exist."
            print(msg)
            return False , msg
        else:
            return True, ""
    if sheet_name not in wb.sheetnames:
        msg = f"Product sheet '{sheet_name}' does not exist."
        print(msg)
        return False , msg
    return True , ""


def add_product( file_name: str , name: str , description: str , stock: int , price: int ) -> Union[bool , str]:
    """Add a new product with the current date."""
    sheet_name = name.lower()
    is_valid , msg = validate_sheet_exists(file_name , sheet_name, flag = True)
    if not is_valid:
        return False , msg

    wb = prepare_workbook(file_name)
    ws = wb.create_sheet(title = sheet_name , index = 0)  # Add the sheet as the first sheet
    ws.append(["Transaction Date" , "Name" , "Description" , "Stock" , "Price"])

    # Apply header styles and data validation for the product sheet
    apply_header_styles(ws)
    apply_data_validation(ws)

    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S") , name , description , stock , price])
    save_changes(wb , file_name)
    msg = f"Product sheet '{sheet_name}' added successfully."
    print(msg)
    return True , msg


def edit_product( file_name: str , current_sheet_index:int , name = None , description = None , stock = None , price = None ) -> Union[bool , str]:
    """Edit an existing product by adding a new record with updated data."""
    wb = prepare_workbook(file_name)
    ws = wb.worksheets[current_sheet_index]
    
    sheet_name = ws.title
    is_valid , msg = validate_sheet_exists(file_name , sheet_name)
    if not is_valid:
        return False , msg
    

    # Retrieve the last row's values in case if given parameters were null
    last_row = ws.max_row
    last_record = {ws.cell(row = 1 , column = col).value: ws.cell(row = last_row , column = col).value for col in range(1 , ws.max_column + 1)}

    # Prepare the new record with existing values and update with provided arguments
    new_record = {
        "Transaction Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S") ,
        "Name": name if name is not None else last_record["Name"] ,
        "Description": description if description is not None else last_record["Description"] ,
        "Stock": stock if stock is not None else last_record["Stock"] ,
        "Price": price if price is not None else last_record["Price"]
    }

    ws.append([new_record["Transaction Date"] , new_record["Name"] , new_record["Description"] , new_record["Stock"] , new_record['Price']])
    if name and name.lower() != last_record['Name'].lower():  # Check for case-insensitive name change
        ws.title = name
    save_changes(wb , file_name)
    msg = f"Product sheet '{sheet_name}' updated successfully."
    print(msg)
    return True , msg


def delete_product_sheet( file_name: str , sheet_name: str ) -> Union[bool , str]:
    """Delete a product sheet."""
    is_valid , msg = validate_sheet_exists(file_name , sheet_name)
    if not is_valid:
        return False , msg

    wb = prepare_workbook(file_name)
    del wb[sheet_name]
    save_changes(wb , file_name)
    msg = f"Product sheet '{sheet_name}' deleted successfully."
    print(msg)
    return True , msg

def delete_last_row(file_name: str , sheet_index: int ) -> Union[bool, str]:
    "Delete the last row of data from given index."
    try:
        wb = prepare_workbook(file_name)
        sheet_names = wb.sheetnames

        if sheet_index < 0:
            return False, "Invalid Sheet Index."
        
        ws = wb[sheet_names[sheet_index]]

        last_row = ws.max_row

        # Validation for header 
        if last_row > 1:
            # To update the title of worksheet
            new_title = ws.cell(row=last_row - 1, column=1).value
            if new_title:
                ws.title = str(new_title)
            ws.delete_rows(last_row)
            save_changes(wb, file_name)
            return True, f"Last row of {sheet_names[sheet_index]} deleted successfully."
        else:
            return False, "Cannot delete the header."
        
    except Exception as e:
        return False, f"Failed to delete last row: {e}"

# Example usage
create_excel_file()
#! 1
add_product(excel_file_path,  "ProductA Name", "Initial stock", 100, 2000)
add_product(excel_file_path,"ProductMilad Name", "Initial stock", 100, 200)
#? 2
# edit_product(excel_file_path,1, "test12", description="Stock reduced", stock=80, price = 1551)
# edit_product(excel_file_path,1, name="Updated ProductA Name", stock=60)
# edit_product(excel_file_path,1, "ProductA", description="Final update")
# delete_product_sheet(excel_file_path, "test12")
